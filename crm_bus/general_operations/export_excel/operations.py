from datetime import datetime
import os
import pathlib

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

from general_operations.export_excel import consts
import response_codes


class MakeExcelData(object):
    """To create an Excel with the tables data

    Run with the endpoints to export the data in to Excel file.
    """
    def __init__(self, data: list, title: str, special_value: str):
        self.data = data
        self.title = title
        self.special_value = special_value

    def traduce_key(self, key):
        if key in consts.DICT_HEADERS[self.title]:
            return consts.DICT_HEADERS[self.title][key]
        else:
            return key

    def col_title_process(self):
        """Take the key of the one data and add in col_title_list"""
        col_title_list = []
        dict_value_list = []

        for one_data in self.data:
            for key, value in one_data.items():
                if key not in consts.EXCLUDE_HEADER:
                    if self.traduce_key(key) not in col_title_list:
                        col_title_list.append(self.traduce_key(key))
                    if isinstance(value, dict) and self.traduce_key(key) not\
                            in dict_value_list:
                        dict_value_list.append(self.traduce_key(key))

        return col_title_list, dict_value_list

    def row_data(self, col_title: list, dict_value: list):
        """take the value of the data list and add in to a list of list.

        return the list of list with the value data.

        Attributes:
            col_title: list with the key values.
            dict_value: list with the col_title have dict has a value.
        """
        list_row = []
        total_row = []
        for each_data in self.data:
            try:
                for key, value in each_data.items():
                    if self.traduce_key(key) in col_title:
                        if value:
                            if self.traduce_key(key) in dict_value:
                                list_row.append(value[self.special_value])
                            else:
                                list_row.append(value)
                        else:
                            list_row.append(consts.NOT_INFORMATION)
                    elif key not in consts.EXCLUDE_HEADER:
                        list_row.append(consts.NOT_INFORMATION)
                total_row.append(list_row)
                list_row = []
            except AssertionError as ax:
                pass

        return total_row

    @staticmethod
    def styles_cells(sheet, num_rows: int, num_col_title: int):
        thin = Side(border_style="thin", color="000000")
        for idx in range(num_col_title):

            # header
            header = sheet.cell(row=4, column=(idx + 1))
            header.fill = PatternFill(fgColor="B1C6F8", fill_type="solid")
            header.alignment = Alignment(horizontal='center', vertical='center')
            header.font = Font(name='Arial', size=14, bold=True)
            header.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            column_name = chr(65 + idx)
            sheet.row_dimensions[idx + 1].height = 30
            sheet.column_dimensions[column_name].width = 30

            for cell_row in range(num_rows):
                # Rows data
                rows = sheet.cell(row=cell_row + 5, column=(idx + 1))
                rows.font = Font(name='Arial', size=12)
                rows.alignment = Alignment(horizontal='center', vertical='center')
                rows.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                sheet.row_dimensions[cell_row + 6].height = 30

    def report_name(self):
        list_name = []
        date_now = datetime.strftime(datetime.now(), '%d-%m-%Y')
        for each in self.title:
            if each == ' ':
                list_name.append('_')
            else:
                list_name.append(each)
        title = "".join(list_name)
        name = title + '_' + date_now + '.xls'
        return_filename = os.path.join('crm_bus/template/', name)
        return return_filename

    def excel_generator(self):
        filename = None
        date_now = datetime.strftime(datetime.now(), '%d-%m-%Y')
        try:
            # Create the file
            wb = openpyxl.Workbook()
            sheet = wb.active
            col_title, dict_value = self.col_title_process()
            tota_row = self.row_data(col_title, dict_value)
            numb_col = len(col_title)
            # paint the title and date

            # Tittle
            if numb_col % 2 != 0:
                numb_col += 1
                numb_col = numb_col / 2
            else:
                numb_col = numb_col / 2
            sheet.merge_cells(start_row=1, end_row=1, start_column=(numb_col - 1),
                              end_column=(numb_col + 3))
            title = sheet.cell(row=1, column=(numb_col - 1), value=self.title)
            title.font = Font(name='Arial', size=14, bold=True)
            title.alignment = Alignment(horizontal='center')

            # Date
            date = sheet.cell(row=3, column=1, value='Create:')
            date.font = Font(name='Arial', size=14, bold=True)
            date_data = sheet.cell(row=3, column=2, value=date_now)
            date_data.font = Font(name='Arial', size=14, bold=True)

            # paint the col_tittle
            sheet.append(col_title)

            # paint the row data
            for row in tota_row:
                sheet.append(row)

            # styles of cells
            self.styles_cells(sheet, num_rows=len(tota_row), num_col_title=len(col_title))

            # save
            filename = self.report_name()
            wb.save(filename)

            return_code = response_codes.SUCCESS
        except Exception as ex:
            os.system(f'Error in excel_generator{ex}')
            return_code = response_codes.UNEXPECTED_ERROR

        return return_code, filename

    def delete_excel(self, max_files):
        initial_count = 0
        for path in pathlib.Path('crm_bus/template/').iterdir():
            if path.is_file():
                initial_count += 1

        if initial_count > max_files:
            files = os.listdir('crm_bus/template/')
            for name in files:
                os.remove(os.path.join('crm_bus/template/', name))
